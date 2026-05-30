"""Stundenkonto: Zeitbuchungen aus Tagesberichten (Welle Z1, Backend-only)."""

from __future__ import annotations

import json
import re
import uuid
from datetime import date, datetime, timezone
from io import StringIO
import csv
from pathlib import Path
from typing import Any, Callable

DEFAULT_BREAK_MINUTES = 45
_ENTRIES_FILE: Path | None = None


def configure(entries_file: Path) -> None:
    global _ENTRIES_FILE
    _ENTRIES_FILE = entries_file


def _entries_path() -> Path:
    if _ENTRIES_FILE is None:
        raise RuntimeError("time_account.configure() was not called")
    return _ENTRIES_FILE


def _read_entries_doc(read_json: Callable[..., Any]) -> dict[str, Any]:
    return read_json(_entries_path(), {"entries": []})


def _write_entries_doc(write_json: Callable[..., None], doc: dict[str, Any]) -> None:
    write_json(_entries_path(), doc)


def _normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip()).casefold()


def parse_hhmm(value: str) -> int | None:
    s = str(value or "").strip()
    m = re.match(r"^(\d{1,2}):(\d{2})$", s)
    if not m:
        return None
    hour, minute = int(m.group(1)), int(m.group(2))
    if hour > 23 or minute > 59:
        return None
    return hour * 60 + minute


def compute_booked_hours(start_time: str, end_time: str, break_minutes: int) -> float | None:
    start_min = parse_hhmm(start_time)
    end_min = parse_hhmm(end_time)
    if start_min is None or end_min is None:
        return None
    if end_min <= start_min:
        return None
    gross = end_min - start_min
    pause = max(0, int(break_minutes))
    net = gross - pause
    if net <= 0:
        return 0.0
    return round(net / 60.0, 2)


def _parse_report_date(date_raw: Any) -> date | None:
    s = str(date_raw or "").strip()
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def _iso_week_key(d: date) -> tuple[int, int]:
    iso = d.isocalendar()
    return iso[0], iso[1]


def match_employees_by_names(names: list[str], employees: list[dict[str, Any]]) -> list[tuple[dict[str, Any], str]]:
    index: dict[str, dict[str, Any]] = {}
    for emp in employees:
        if not isinstance(emp, dict):
            continue
        key = _normalize_name(str(emp.get("name") or ""))
        if key and key not in index:
            index[key] = emp

    matched: list[tuple[dict[str, Any], str]] = []
    seen_ids: set[str] = set()
    for raw in names:
        label = str(raw or "").strip()
        if not label:
            continue
        emp = index.get(_normalize_name(label))
        if emp is None:
            continue
        emp_id = str(emp.get("id") or "")
        if not emp_id or emp_id in seen_ids:
            continue
        seen_ids.add(emp_id)
        matched.append((emp, label))
    return matched


def match_employees_by_ids(
    employee_ids: list[str],
    employees: list[dict[str, Any]],
) -> list[tuple[dict[str, Any], str]]:
    by_id: dict[str, dict[str, Any]] = {}
    for emp in employees:
        if not isinstance(emp, dict):
            continue
        emp_id = str(emp.get("id") or "").strip()
        if emp_id:
            by_id[emp_id] = emp

    matched: list[tuple[dict[str, Any], str]] = []
    seen_ids: set[str] = set()
    for raw_id in employee_ids:
        emp_id = str(raw_id or "").strip()
        if not emp_id or emp_id in seen_ids:
            continue
        emp = by_id.get(emp_id)
        if emp is None:
            continue
        seen_ids.add(emp_id)
        matched.append((emp, str(emp.get("name") or "")))
    return matched


def resolve_report_employees(
    report: dict[str, Any],
    employees: list[dict[str, Any]],
) -> tuple[list[tuple[dict[str, Any], str]], list[str], str]:
    """Bevorzugt employeeIds, faellt auf Namen zurueck."""
    raw_ids = report.get("employeeIds")
    if isinstance(raw_ids, list):
        ids = [str(x).strip() for x in raw_ids if str(x).strip()]
        if ids:
            matched = match_employees_by_ids(ids, employees)
            skipped_ids = [i for i in ids if i not in {str(e.get("id") or "") for e, _ in matched}]
            return matched, skipped_ids, "ids"

    raw_names = report.get("employees")
    names = [str(n) for n in raw_names] if isinstance(raw_names, list) else []
    matched = match_employees_by_names(names, employees)
    matched_norm = {_normalize_name(label) for _, label in matched}
    skipped_names = [n for n in names if _normalize_name(n) and _normalize_name(n) not in matched_norm]
    return matched, skipped_names, "names"


def _entry_source(entry: dict[str, Any]) -> str:
    src = str(entry.get("source") or "").strip()
    if src in ("report", "manual"):
        return src
    return "report" if entry.get("reportId") else "manual"


def _entry_api_item(entry: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": entry.get("id"),
        "source": _entry_source(entry),
        "reportId": entry.get("reportId"),
        "employeeId": entry.get("employeeId"),
        "employeeName": entry.get("employeeName"),
        "date": entry.get("date"),
        "projectId": entry.get("projectId"),
        "projectName": entry.get("projectName"),
        "startTime": entry.get("startTime"),
        "endTime": entry.get("endTime"),
        "breakMinutes": entry.get("breakMinutes"),
        "hours": entry.get("hours"),
        "note": entry.get("note") or "",
        "createdAt": entry.get("createdAt"),
        "updatedAt": entry.get("updatedAt"),
    }


def list_time_entries(
    *,
    read_json: Callable[..., Any],
    employee_id: str | None = None,
    month: str | None = None,
) -> list[dict[str, Any]]:
    entries = list(_read_entries_doc(read_json).get("entries") or [])
    out: list[dict[str, Any]] = []
    month_prefix = str(month or "").strip()[:7] if month else ""

    for entry in entries:
        if not isinstance(entry, dict):
            continue
        if employee_id and str(entry.get("employeeId") or "") != employee_id:
            continue
        if month_prefix and not str(entry.get("date") or "").startswith(month_prefix):
            continue
        out.append(_entry_api_item(entry))

    out.sort(key=lambda e: (str(e.get("date") or ""), str(e.get("createdAt") or "")), reverse=True)
    return out


def delete_time_entry(
    entry_id: str,
    *,
    read_json: Callable[..., Any],
    write_json: Callable[..., None],
) -> bool:
    eid = str(entry_id or "").strip()
    if not eid:
        return False
    doc = _read_entries_doc(read_json)
    kept: list[dict[str, Any]] = []
    removed = False
    for entry in doc.get("entries") or []:
        if isinstance(entry, dict) and str(entry.get("id") or "") == eid:
            removed = True
            continue
        kept.append(entry)
    if removed:
        doc["entries"] = kept
        _write_entries_doc(write_json, doc)
    return removed


def _find_employee(employees: list[dict[str, Any]], employee_id: str) -> dict[str, Any] | None:
    emp_id = str(employee_id or "").strip()
    if not emp_id:
        return None
    for emp in employees:
        if isinstance(emp, dict) and str(emp.get("id") or "") == emp_id:
            return emp
    return None


def create_manual_time_entry(
    *,
    employee_id: str,
    date_str: str,
    hours: float,
    note: str,
    employees: list[dict[str, Any]],
    read_json: Callable[..., Any],
    write_json: Callable[..., None],
) -> dict[str, Any]:
    emp = _find_employee(employees, employee_id)
    if emp is None:
        raise ValueError("employee_not_found")

    entry_date = _parse_report_date(date_str)
    if entry_date is None:
        raise ValueError("invalid_date")

    booked_hours = round(float(hours), 2)
    if booked_hours == 0:
        raise ValueError("hours_zero")
    if booked_hours < -24 or booked_hours > 24:
        raise ValueError("hours_out_of_range")

    note_clean = re.sub(r"\s+", " ", (note or "").strip())
    if len(note_clean) < 2:
        raise ValueError("note_required")

    now = datetime.now(timezone.utc).isoformat()
    entry = {
        "id": str(uuid.uuid4()),
        "source": "manual",
        "reportId": None,
        "employeeId": str(emp.get("id") or ""),
        "employeeName": str(emp.get("name") or ""),
        "date": entry_date.isoformat(),
        "projectId": None,
        "projectName": "Manuelle Korrektur",
        "startTime": "",
        "endTime": "",
        "breakMinutes": 0,
        "hours": booked_hours,
        "note": note_clean,
        "createdAt": now,
        "updatedAt": now,
    }

    doc = _read_entries_doc(read_json)
    entries = list(doc.get("entries") or [])
    entries.append(entry)
    doc["entries"] = entries
    _write_entries_doc(write_json, doc)
    return _entry_api_item(entry)


def delete_entries_for_report(
    report_id: str,
    *,
    read_json: Callable[..., Any],
    write_json: Callable[..., None],
) -> int:
    rid = str(report_id or "").strip()
    if not rid:
        return 0
    doc = _read_entries_doc(read_json)
    kept: list[dict[str, Any]] = []
    removed = 0
    for entry in doc.get("entries") or []:
        if isinstance(entry, dict) and str(entry.get("reportId") or "") == rid:
            removed += 1
            continue
        kept.append(entry)
    if removed:
        doc["entries"] = kept
        _write_entries_doc(write_json, doc)
    return removed


def sync_entries_for_report(
    report: dict[str, Any],
    employees: list[dict[str, Any]],
    *,
    read_json: Callable[..., Any],
    write_json: Callable[..., None],
) -> dict[str, Any]:
    """Erzeugt Zeitbuchungen fuer einen gespeicherten Bericht (idempotent pro reportId)."""
    report_id = str(report.get("id") or "").strip()
    if not report_id:
        return {
            "created": 0,
            "hoursPerEmployee": None,
            "skippedNames": [],
            "skippedEmployeeIds": [],
            "reason": "missing_report_id",
        }

    start_time = str(report.get("startTime") or "")
    end_time = str(report.get("endTime") or "")
    break_minutes = int(report.get("breakMinutes") if report.get("breakMinutes") is not None else DEFAULT_BREAK_MINUTES)
    break_minutes = max(0, min(break_minutes, 480))

    hours = compute_booked_hours(start_time, end_time, break_minutes)
    if hours is None:
        delete_entries_for_report(report_id, read_json=read_json, write_json=write_json)
        return {
            "created": 0,
            "hoursPerEmployee": None,
            "skippedNames": [],
            "skippedEmployeeIds": [],
            "reason": "invalid_work_time",
        }

    raw_names = report.get("employees")
    names = [str(n) for n in raw_names] if isinstance(raw_names, list) else []
    matched, skipped, match_mode = resolve_report_employees(report, employees)

    all_names_norm = {_normalize_name(n) for n in names if _normalize_name(n)}
    skipped_names = skipped if match_mode == "names" else []

    delete_entries_for_report(report_id, read_json=read_json, write_json=write_json)

    if not matched:
        reason = "no_matched_employees"
        if match_mode == "ids" and skipped:
            reason = "unknown_employee_ids"
        elif not all_names_norm and match_mode == "names":
            reason = "no_employees"
        return {
            "created": 0,
            "hoursPerEmployee": hours,
            "skippedNames": skipped_names,
            "skippedEmployeeIds": skipped if match_mode == "ids" else [],
            "reason": reason,
        }

    now = datetime.now(timezone.utc).isoformat()
    doc = _read_entries_doc(read_json)
    entries = list(doc.get("entries") or [])

    for emp, label in matched:
        entry = {
            "id": str(uuid.uuid4()),
            "source": "report",
            "reportId": report_id,
            "employeeId": str(emp.get("id") or ""),
            "employeeName": str(emp.get("name") or label),
            "date": str(report.get("date") or ""),
            "projectId": report.get("projectId"),
            "projectName": str(report.get("projectName") or ""),
            "startTime": start_time,
            "endTime": end_time,
            "breakMinutes": break_minutes,
            "hours": hours,
            "createdAt": now,
            "updatedAt": now,
        }
        entries.append(entry)

    doc["entries"] = entries
    _write_entries_doc(write_json, doc)

    return {
        "created": len(matched),
        "hoursPerEmployee": hours,
        "skippedNames": skipped_names,
        "skippedEmployeeIds": [],
        "reason": None,
    }


def _employee_hours_balance_start(employee: dict[str, Any]) -> float:
    raw = employee.get("hoursBalanceStart")
    try:
        return round(float(raw), 2)
    except (TypeError, ValueError):
        return 0.0


def employee_hours_balance_start_date(employee: dict[str, Any]) -> str | None:
    d = _parse_report_date(employee.get("hoursBalanceStartDate"))
    return d.isoformat() if d else None


def build_time_accounts(
    employees: list[dict[str, Any]],
    entries: list[dict[str, Any]],
    *,
    month: str | None = None,
    today: date | None = None,
) -> dict[str, Any]:
    ref_day = today or date.today()
    week_key = _iso_week_key(ref_day)
    month_prefix = str(month or ref_day.strftime("%Y-%m")).strip()[:7]

    booked_total: dict[str, float] = {}
    week_total: dict[str, float] = {}
    month_total: dict[str, float] = {}
    entry_count: dict[str, int] = {}

    for entry in entries:
        if not isinstance(entry, dict):
            continue
        emp_id = str(entry.get("employeeId") or "")
        if not emp_id:
            continue
        try:
            hours = float(entry.get("hours") or 0)
        except (TypeError, ValueError):
            hours = 0.0
        booked_total[emp_id] = round(booked_total.get(emp_id, 0.0) + hours, 2)
        entry_count[emp_id] = entry_count.get(emp_id, 0) + 1

        entry_date = _parse_report_date(entry.get("date"))
        if entry_date and _iso_week_key(entry_date) == week_key:
            week_total[emp_id] = round(week_total.get(emp_id, 0.0) + hours, 2)
        if month_prefix and str(entry.get("date") or "").startswith(month_prefix):
            month_total[emp_id] = round(month_total.get(emp_id, 0.0) + hours, 2)

    accounts: list[dict[str, Any]] = []
    for emp in employees:
        if not isinstance(emp, dict):
            continue
        emp_id = str(emp.get("id") or "")
        if not emp_id:
            continue
        start_balance = _employee_hours_balance_start(emp)
        start_date = employee_hours_balance_start_date(emp)
        booked = booked_total.get(emp_id, 0.0)
        accounts.append(
            {
                "employeeId": emp_id,
                "employeeName": str(emp.get("name") or ""),
                "active": bool(emp.get("active", True)),
                "hoursBalanceStart": start_balance,
                "hoursBalanceStartDate": start_date,
                "bookedHoursTotal": booked,
                "currentBalance": round(start_balance + booked, 2),
                "weekHours": week_total.get(emp_id, 0.0),
                "monthHours": month_total.get(emp_id, 0.0),
                "entryCount": entry_count.get(emp_id, 0),
            }
        )

    accounts.sort(key=lambda a: str(a.get("employeeName") or "").casefold())
    return {"accounts": accounts, "month": month_prefix}


def list_time_accounts(
    employees: list[dict[str, Any]],
    *,
    read_json: Callable[..., Any],
    month: str | None = None,
    today: date | None = None,
) -> dict[str, Any]:
    entries = list(_read_entries_doc(read_json).get("entries") or [])
    return build_time_accounts(employees, entries, month=month, today=today)


def _format_de_decimal(value: float) -> str:
    return f"{value:.2f}".replace(".", ",")


def _format_export_date(date_raw: Any) -> str:
    d = _parse_report_date(date_raw)
    if d is None:
        return str(date_raw or "")
    return d.strftime("%d.%m.%Y")


def _entry_type_label(entry: dict[str, Any]) -> str:
    return "Korrektur" if _entry_source(entry) == "manual" else "Bericht"


def _sorted_export_entries(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        entries,
        key=lambda e: (
            str(e.get("employeeName") or "").casefold(),
            str(e.get("date") or ""),
            str(e.get("createdAt") or ""),
        ),
    )


def _export_summary_accounts(accounts: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = [
        a
        for a in accounts
        if float(a.get("monthHours") or 0) != 0
        or float(a.get("hoursBalanceStart") or 0) != 0
        or float(a.get("bookedHoursTotal") or 0) != 0
    ]
    rows.sort(key=lambda a: str(a.get("employeeName") or "").casefold())
    return rows


def _entry_hours_value(entry: dict[str, Any]) -> float:
    try:
        return float(entry.get("hours") or 0)
    except (TypeError, ValueError):
        return 0.0


def _build_report_label(entry: dict[str, Any], report: dict[str, Any] | None = None) -> str:
    if _entry_source(entry) == "manual":
        return ""
    rep = report if isinstance(report, dict) else {}
    parts: list[str] = []
    date_raw = rep.get("date") or entry.get("date")
    date_label = _format_export_date(date_raw)
    if date_label:
        parts.append(date_label)
    project = str(rep.get("projectName") or entry.get("projectName") or "").strip()
    if project:
        parts.append(project)
    customer = str(rep.get("customerName") or "").strip()
    if customer and customer.lower() not in ("keine angabe",) and customer != project:
        parts.append(customer)
    return " · ".join(parts)


def enrich_entries_for_export(
    entries: list[dict[str, Any]],
    reports_by_id: dict[str, dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    index = reports_by_id or {}
    out: list[dict[str, Any]] = []
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        item = dict(entry)
        rid = str(item.get("reportId") or "").strip()
        item["reportLabel"] = _build_report_label(item, index.get(rid))
        out.append(item)
    return out


def _apply_excel_column_widths(
    ws: Any,
    *,
    min_col: int,
    max_col: int,
    first_row: int,
    last_row: int,
    min_widths: dict[int, float] | None = None,
    max_widths: dict[int, float] | None = None,
) -> None:
    from openpyxl.utils import get_column_letter

    mins = min_widths or {}
    maxs = max_widths or {}
    for col in range(min_col, max_col + 1):
        width = float(mins.get(col, 10))
        cap = float(maxs.get(col, 56))
        for row in range(first_row, last_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            text = str(value)
            width = max(width, len(text) * 1.08 + 2.5)
        ws.column_dimensions[get_column_letter(col)].width = min(width, cap)


def build_time_export_csv(
    *,
    entries: list[dict[str, Any]],
    accounts: list[dict[str, Any]],
    month: str,
    company_name: str = "",
) -> bytes:
    """CSV fuer Buchhaltung (Semikolon, UTF-8 mit BOM, Dezimal komma)."""
    month_prefix = str(month or "").strip()[:7]
    buf = StringIO()
    writer = csv.writer(buf, delimiter=";", lineterminator="\r\n")

    if company_name.strip():
        writer.writerow(["Firma", company_name.strip()])
    writer.writerow(["Monat", month_prefix])
    writer.writerow([])

    writer.writerow(
        [
            "Mitarbeiter",
            "Datum",
            "Baustelle",
            "Von",
            "Bis",
            "Pause_Min",
            "Stunden",
            "Typ",
            "Grund",
            "Bericht",
        ]
    )

    detail_rows = _sorted_export_entries(entries)

    for entry in detail_rows:
        hours = _entry_hours_value(entry)
        writer.writerow(
            [
                str(entry.get("employeeName") or ""),
                _format_export_date(entry.get("date")),
                str(entry.get("projectName") or ""),
                str(entry.get("startTime") or ""),
                str(entry.get("endTime") or ""),
                str(entry.get("breakMinutes") or ""),
                _format_de_decimal(hours),
                _entry_type_label(entry),
                str(entry.get("note") or ""),
                str(entry.get("reportLabel") or ""),
            ]
        )

    writer.writerow([])
    writer.writerow(["ZUSAMMENFASSUNG"])
    writer.writerow(
        [
            "Mitarbeiter",
            "Monat",
            "Summe_Monat",
            "Startsaldo",
            "Stand_Startsaldo",
            "Gesamt_gebucht",
            "Aktueller_Saldo",
        ]
    )

    for acct in _export_summary_accounts(accounts):
        start_date = acct.get("hoursBalanceStartDate")
        writer.writerow(
            [
                str(acct.get("employeeName") or ""),
                month_prefix,
                _format_de_decimal(float(acct.get("monthHours") or 0)),
                _format_de_decimal(float(acct.get("hoursBalanceStart") or 0)),
                _format_export_date(start_date) if start_date else "",
                _format_de_decimal(float(acct.get("bookedHoursTotal") or 0)),
                _format_de_decimal(float(acct.get("currentBalance") or 0)),
            ]
        )

    return buf.getvalue().encode("utf-8-sig")


def build_time_export_xlsx(
    *,
    entries: list[dict[str, Any]],
    accounts: list[dict[str, Any]],
    month: str,
    company_name: str = "",
) -> bytes:
    """Excel-Export fuer Buchhaltung (formatierte .xlsx)."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    month_prefix = str(month or "").strip()[:7]
    wb = Workbook()
    ws = wb.active
    ws.title = "Stundenkonto"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    section_font = Font(bold=True, size=12)
    hours_fmt = "#,##0.00"
    text_align = Alignment(vertical="top", wrap_text=False)

    row_idx = 1
    if company_name.strip():
        ws.cell(row=row_idx, column=1, value="Firma").font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=company_name.strip())
        row_idx += 1
    ws.cell(row=row_idx, column=1, value="Monat").font = Font(bold=True)
    ws.cell(row=row_idx, column=2, value=month_prefix)
    row_idx += 2

    detail_headers = [
        "Mitarbeiter",
        "Datum",
        "Baustelle",
        "Von",
        "Bis",
        "Pause (Min)",
        "Stunden",
        "Typ",
        "Grund",
        "Bericht",
    ]
    for col, label in enumerate(detail_headers, start=1):
        cell = ws.cell(row=row_idx, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    detail_header_row = row_idx
    row_idx += 1

    for entry in _sorted_export_entries(entries):
        hours = _entry_hours_value(entry)
        pause_raw = entry.get("breakMinutes")
        try:
            pause_val = int(pause_raw) if pause_raw not in (None, "") else None
        except (TypeError, ValueError):
            pause_val = None
        ws.cell(row=row_idx, column=1, value=str(entry.get("employeeName") or "")).alignment = text_align
        ws.cell(row=row_idx, column=2, value=_format_export_date(entry.get("date"))).alignment = text_align
        ws.cell(row=row_idx, column=3, value=str(entry.get("projectName") or "")).alignment = text_align
        ws.cell(row=row_idx, column=4, value=str(entry.get("startTime") or "")).alignment = text_align
        ws.cell(row=row_idx, column=5, value=str(entry.get("endTime") or "")).alignment = text_align
        if pause_val is not None:
            ws.cell(row=row_idx, column=6, value=pause_val).alignment = text_align
        hours_cell = ws.cell(row=row_idx, column=7, value=hours)
        hours_cell.number_format = hours_fmt
        hours_cell.alignment = text_align
        ws.cell(row=row_idx, column=8, value=_entry_type_label(entry)).alignment = text_align
        ws.cell(row=row_idx, column=9, value=str(entry.get("note") or "")).alignment = text_align
        ws.cell(row=row_idx, column=10, value=str(entry.get("reportLabel") or "")).alignment = text_align
        row_idx += 1

    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Zusammenfassung").font = section_font
    row_idx += 1

    summary_headers = [
        "Mitarbeiter",
        "Monat",
        "Summe Monat",
        "Startsaldo",
        "Stand Startsaldo",
        "Gesamt gebucht",
        "Aktueller Saldo",
    ]
    for col, label in enumerate(summary_headers, start=1):
        cell = ws.cell(row=row_idx, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    row_idx += 1

    for acct in _export_summary_accounts(accounts):
        start_date = acct.get("hoursBalanceStartDate")
        ws.cell(row=row_idx, column=1, value=str(acct.get("employeeName") or ""))
        ws.cell(row=row_idx, column=2, value=month_prefix)
        for col, key in ((3, "monthHours"), (4, "hoursBalanceStart"), (6, "bookedHoursTotal"), (7, "currentBalance")):
            cell = ws.cell(row=row_idx, column=col, value=float(acct.get(key) or 0))
            cell.number_format = hours_fmt
        ws.cell(
            row=row_idx,
            column=5,
            value=_format_export_date(start_date) if start_date else "",
        )
        row_idx += 1

    last_row = row_idx - 1
    _apply_excel_column_widths(
        ws,
        min_col=1,
        max_col=10,
        first_row=1,
        last_row=last_row,
        min_widths={
            1: 14,
            2: 12,
            3: 18,
            4: 7,
            5: 7,
            6: 12,
            7: 10,
            8: 11,
            9: 20,
            10: 38,
        },
        max_widths={
            1: 24,
            2: 14,
            3: 32,
            4: 9,
            5: 9,
            6: 14,
            7: 12,
            8: 14,
            9: 40,
            10: 58,
        },
    )

    ws.freeze_panes = ws.cell(row=detail_header_row + 1, column=1)
    last_detail_row = detail_header_row + len(_sorted_export_entries(entries))
    if last_detail_row > detail_header_row:
        ws.auto_filter.ref = f"A{detail_header_row}:J{last_detail_row}"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
